"""
PDF Click Counter
-----------------
A desktop tool for counting items on PDF building plans (or any PDF).

Usage:
    python pdf_click_counter.py

Requirements:
    pip install pymupdf Pillow

Features:
- Load any PDF file via file dialog
- Navigate multi-page PDFs
- Click to place numbered markers
- Multiple named categories with unique colours
- Semi-transparent markers (alpha 0.6) rendered via Pillow
- Minimum distance guard (won't count same spot twice)
- Undo last click (per category)
- Reset page counts (per category)
- Zoom in/out — markers stay anchored to PDF coordinates
- Export a summary of counts per category per page
"""

import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser, simpledialog, ttk
import math
import json
import os

try:
    import fitz  # PyMuPDF
except ImportError:
    raise SystemExit(
        "PyMuPDF is required. Install it with:\n    pip install pymupdf"
    )

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
except ImportError:
    raise SystemExit(
        "Pillow is required. Install it with:\n    pip install Pillow"
    )

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False

# ── Configuration ────────────────────────────────────────────────────────────

MIN_DISTANCE_PX = 10
DEFAULT_ZOOM    = 1.5
MARKER_RADIUS   = 10     # doubled from original 10
MARKER_ALPHA    = int(0.6 * 255)  # 153

# ── Main Application ──────────────────────────────────────────────────────────

class PDFClickCounter:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("PDF Click Counter")
        self.root.geometry("1100x820")
        self.root.configure(bg="#1e1e2e")

        # PDF state
        self.pdf_doc       = None
        self.pdf_path      = ""
        self.current_page  = 0
        self.zoom          = DEFAULT_ZOOM
        self.tk_image      = None
        self.page_offset_x = 0
        self.page_offset_y = 0

        # Categories: list of {"name": str, "color": "#rrggbb"}
        self.categories: list[dict] = [{"name": "Category 1", "color": "#e63946"}]
        self.current_category: int = 0

        # clicks[cat_idx][page_idx] = [(pdf_x, pdf_y), ...]  (PDF coords, zoom-independent)
        self.clicks: dict[int, dict[int, list[tuple[float, float]]]] = {0: {}}

        # Must keep PhotoImage references alive or tkinter GC's them
        self._marker_images: list = []

        self._pan_key_held: bool = False

        # Ruler state
        self.ruler_mode: bool = False
        # 0–2 points in PDF page coords; third click starts a fresh measurement
        self.ruler_points: list[tuple[float, float]] = []
        # metres per PDF point; None = show raw PDF units
        self.scale_m_per_pt: float | None = None

        self._build_ui()
        self._bind_keys()

    # ── UI Construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        btn_cfg = dict(bg="#313244", fg="#a6adc8", activebackground="#45475a",
                       activeforeground="#2569F2", relief=tk.FLAT,
                       padx=8, pady=4, cursor="hand2", font=("Courier", 10))
 
        # ── Wrapper holds the main bar + the collapsible ruler bar ───────────
        # Keeping both inside one wrapper means pack_forget/pack on ruler_bar
        # never disturbs the canvas position below.
        tools_wrapper = tk.Frame(self.root, bg="#181825")
        tools_wrapper.pack(side=tk.TOP, fill=tk.X)

        # ── Single combined toolbar ──────────────────────────────────────────
        toolbar = tk.Frame(tools_wrapper, bg="#181825", pady=5, padx=8)
        toolbar.pack(side=tk.TOP, fill=tk.X)

        # Page nav – pack RIGHT first so it always hugs the right edge
        nav_frame = tk.Frame(toolbar, bg="#181825")
        nav_frame.pack(side=tk.RIGHT, padx=4)
        tk.Button(nav_frame, text="◀", command=self.prev_page, **btn_cfg).pack(side=tk.LEFT, padx=2)
        self.page_label = tk.Label(nav_frame, text="—", bg="#181825",
                                   fg="#313244", font=("Courier", 10), width=12)
        self.page_label.pack(side=tk.LEFT, padx=4)
        tk.Button(nav_frame, text="▶", command=self.next_page, **btn_cfg).pack(side=tk.LEFT, padx=2)

        # Hamburger menu (Open / Export)
        self._menu_btn = tk.Button(toolbar, text="☰", command=self._show_file_menu, **btn_cfg)
        self._menu_btn.pack(side=tk.LEFT, padx=2)
        self._file_menu = tk.Menu(self.root, tearoff=0, bg="#313244", fg="#a6adc8",
                                  activebackground="#45475a", activeforeground="#a6adc8",
                                  font=("Courier", 10))
        self._file_menu.add_command(label="📂  Open PDF",       command=self.open_pdf)
        self._file_menu.add_separator()
        self._file_menu.add_command(label="💾  Save Session",   command=self.save_session)
        self._file_menu.add_command(label="📂  Load Session",   command=self.load_session)
        self._file_menu.add_separator()
        self._file_menu.add_command(label="📊  Export Summary", command=self.export_summary)

        # Edit actions
        tk.Button(toolbar, text="↩ Undo",  command=self.undo_click, **btn_cfg).pack(side=tk.LEFT, padx=2)
        tk.Button(toolbar, text="🗑 Reset", command=self.reset_page, **btn_cfg).pack(side=tk.LEFT, padx=2)

        tk.Frame(toolbar, bg="#45475a", width=1).pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=3)

        # Zoom
        tk.Label(toolbar, text="Zoom:", bg="#181825", fg="#a6adc8",
                 font=("Courier", 10)).pack(side=tk.LEFT)
        tk.Button(toolbar, text="＋", command=self.zoom_in,  **btn_cfg).pack(side=tk.LEFT, padx=1)
        tk.Button(toolbar, text="－", command=self.zoom_out, **btn_cfg).pack(side=tk.LEFT, padx=1)

        tk.Frame(toolbar, bg="#45475a", width=1).pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=3)

        # Ruler toggle icon (expands ruler bar below when active)
        self.ruler_btn = tk.Button(toolbar, text="📏", command=self.toggle_ruler_mode, **btn_cfg)
        self.ruler_btn.pack(side=tk.LEFT, padx=2)

        tk.Frame(toolbar, bg="#45475a", width=1).pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=3)

        # Category controls
        tk.Label(toolbar, text="Category:", bg="#181825", fg="#a6adc8",
                 font=("Courier", 10)).pack(side=tk.LEFT, padx=(0, 4))

        self.cat_var = tk.StringVar()
        self.cat_menu = ttk.Combobox(toolbar, textvariable=self.cat_var,
                                     state="readonly", font=("Courier", 10), width=18)
        self.cat_menu.pack(side=tk.LEFT, padx=2)
        self.cat_menu.bind("<<ComboboxSelected>>", self._on_category_select)

        self.cat_color_swatch = tk.Label(toolbar, text="  ",
                                         bg=self.categories[0]["color"],
                                         relief=tk.FLAT, width=2)
        self.cat_color_swatch.pack(side=tk.LEFT, padx=(2, 4))

        tk.Button(toolbar, text="＋",        command=self.add_category,         **btn_cfg).pack(side=tk.LEFT, padx=1)
        tk.Button(toolbar, text="✎ Rename", command=self.rename_category,       **btn_cfg).pack(side=tk.LEFT, padx=1)
        tk.Button(toolbar, text="🎨",       command=self.change_category_color, **btn_cfg).pack(side=tk.LEFT, padx=1)

        self._refresh_category_menu()

        # ── Ruler toolbar (collapsible – hidden until ruler mode is on) ──────
        self.ruler_bar = tk.Frame(tools_wrapper, bg="#181825", pady=5, padx=8)
        # Not packed here – toggle_ruler_mode shows/hides it inside tools_wrapper

        rbtn = dict(bg="#313244", fg="#a6adc8", activebackground="#45475a",
                    activeforeground="#2569F2", relief=tk.FLAT,
                    padx=8, pady=3, cursor="hand2", font=("Courier", 10))

        tk.Button(self.ruler_bar, text="⚖ Set Scale", command=self.set_scale,    **rbtn).pack(side=tk.LEFT, padx=4)
        tk.Button(self.ruler_bar, text="✕ Clear",      command=self.clear_ruler,  **rbtn).pack(side=tk.LEFT, padx=4)

        tk.Frame(self.ruler_bar, bg="#45475a", width=1).pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=3)

        tk.Label(self.ruler_bar, text="Scale  1:", bg="#12122a",
                 fg="#a6adc8", font=("Courier", 10)).pack(side=tk.LEFT)
        self.scale_ratio_var = tk.StringVar()
        ratio_entry = tk.Entry(self.ruler_bar, textvariable=self.scale_ratio_var,
                               width=7, bg="#313244", fg="#a6adc8",
                               insertbackground="#a6adc8", relief=tk.FLAT,
                               font=("Courier", 10))
        ratio_entry.pack(side=tk.LEFT, padx=(2, 2))
        ratio_entry.bind("<Return>", lambda _: self.apply_scale_ratio())
        tk.Button(self.ruler_bar, text="Apply", command=self.apply_scale_ratio, **rbtn).pack(side=tk.LEFT, padx=(2, 10))

        self.scale_label = tk.Label(self.ruler_bar, text="Scale: not set",
                                    bg="#181825", fg="#a6adc8", font=("Courier", 10))
        self.scale_label.pack(side=tk.LEFT, padx=(0, 10))

        self.ruler_dist_label = tk.Label(self.ruler_bar, text="",
                                         bg="#181825", fg="#2569F2",
                                         font=("Courier", 11, "bold"))
        self.ruler_dist_label.pack(side=tk.LEFT, padx=4)

        # ── Status bar ───────────────────────────────────────────────────────
        status_bar = tk.Frame(self.root, bg="#11111b", pady=4, padx=10)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        self.status_var = tk.StringVar(value="Open a PDF to begin.")
        tk.Label(status_bar, textvariable=self.status_var,
                 bg="#11111b", fg="#a6adc8",
                 font=("Courier", 10), anchor=tk.W).pack(side=tk.LEFT)

        self.count_var = tk.StringVar(value="Total: 0")
        tk.Label(status_bar, textvariable=self.count_var,
                 bg="#11111b", fg="#a6e3a1",
                 font=("Courier", 13, "bold")).pack(side=tk.RIGHT)

        # ── Scrollable canvas ─────────────────────────────────────────────────
        canvas_frame = tk.Frame(self.root, bg="#1e1e2e")
        canvas_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(canvas_frame, bg="#181825",
                                cursor="crosshair", highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        v_scroll = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL,  command=self.canvas.yview)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        h_scroll = tk.Scrollbar(self.root, orient=tk.HORIZONTAL, command=self.canvas.xview)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        self.canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        self.canvas.bind("<Button-1>",           self.on_canvas_click)
        self.canvas.bind("<B1-Motion>",          self._on_b1_motion)
        self.canvas.bind("<Button-2>",           self._pan_start)
        self.canvas.bind("<B2-Motion>",          self._pan_move)
        self.canvas.bind("<MouseWheel>",         self._on_mousewheel)
        self.canvas.bind("<Shift-MouseWheel>",   self._on_hscroll)
        self.canvas.bind("<Button-4>",           self._on_mousewheel)
        self.canvas.bind("<Button-5>",           self._on_mousewheel)
        self.canvas.bind("<Control-MouseWheel>", self._on_zoom_scroll)
        self.canvas.bind("<Control-Button-4>",   self._on_zoom_scroll)
        self.canvas.bind("<Control-Button-5>",   self._on_zoom_scroll)

    def _bind_keys(self):
        self.root.bind("<Left>",       lambda _: self.prev_page())
        self.root.bind("<Right>",      lambda _: self.next_page())
        self.root.bind("<z>",          lambda _: self.undo_click())
        self.root.bind("<plus>",       lambda _: self.zoom_in())
        self.root.bind("<minus>",      lambda _: self.zoom_out())
        self.root.bind("<KeyPress-d>",   lambda _: self._set_pan_key(True))
        self.root.bind("<KeyRelease-d>", lambda _: self._set_pan_key(False))

    def _show_file_menu(self):
        btn = self._menu_btn
        self._file_menu.tk_popup(btn.winfo_rootx(), btn.winfo_rooty() + btn.winfo_height())

    # ── Category management ───────────────────────────────────────────────────

    def _refresh_category_menu(self):
        names = [c["name"] for c in self.categories]
        self.cat_menu["values"] = names
        self.cat_menu.set(names[self.current_category])
        self.cat_color_swatch.config(bg=self.categories[self.current_category]["color"])

    def _on_category_select(self, *_):
        idx = self.cat_menu.current()
        if idx >= 0:
            self.current_category = idx
            self.cat_color_swatch.config(bg=self.categories[idx]["color"])
            self._update_status()

    def add_category(self):
        name = simpledialog.askstring(
            "New Category",
            "Category name:",
            initialvalue=f"Category {len(self.categories) + 1}",
            parent=self.root,
        )
        if not name:
            return
        result = colorchooser.askcolor(title=f"Choose colour for '{name}'", parent=self.root)
        if not result or not result[1]:
            return
        cat_idx = len(self.categories)
        self.categories.append({"name": name, "color": result[1]})
        self.clicks[cat_idx] = {}
        self.current_category = cat_idx
        self._refresh_category_menu()
        self._update_status()

    def rename_category(self):
        cat = self.categories[self.current_category]
        new_name = simpledialog.askstring(
            "Rename Category", "New name:", initialvalue=cat["name"], parent=self.root
        )
        if new_name:
            cat["name"] = new_name
            self._refresh_category_menu()

    def change_category_color(self):
        cat = self.categories[self.current_category]
        result = colorchooser.askcolor(
            title=f"Choose colour for '{cat['name']}'", color=cat["color"], parent=self.root
        )
        if result and result[1]:
            cat["color"] = result[1]
            self._refresh_category_menu()
            self.render_page()

    # ── PDF Loading & Rendering ───────────────────────────────────────────────

    def open_pdf(self):
        path = filedialog.askopenfilename(
            title="Open PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            self.pdf_doc = fitz.open(path)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not open PDF:\n{exc}")
            return
        self.pdf_path     = path
        self.current_page = 0
        self.clicks       = {i: {} for i in range(len(self.categories))}
        self.root.title(f"PDF Click Counter — {os.path.basename(path)}")
        self.render_page()

    def render_page(self):
        if self.pdf_doc is None:
            return

        page = self.pdf_doc[self.current_page]
        mat  = fitz.Matrix(self.zoom, self.zoom)
        pix  = page.get_pixmap(matrix=mat, alpha=False)

        self.tk_image = tk.PhotoImage(data=pix.tobytes("ppm"))

        canvas_w = self.canvas.winfo_width()  or 900
        canvas_h = self.canvas.winfo_height() or 700
        img_w, img_h = pix.width, pix.height

        self.page_offset_x = max(0, (canvas_w - img_w) // 2)
        self.page_offset_y = 20

        total_w = max(canvas_w, img_w + self.page_offset_x + 20)
        total_h = max(canvas_h, img_h + self.page_offset_y + 20)

        self.canvas.delete("all")
        self.canvas.configure(scrollregion=(0, 0, total_w, total_h))
        self.canvas.create_image(self.page_offset_x, self.page_offset_y,
                                 anchor=tk.NW, image=self.tk_image)

        self._marker_images.clear()

        # Draw all categories on this page so all markers are visible at once
        for cat_idx, cat in enumerate(self.categories):
            cat_clicks = self.clicks.get(cat_idx, {})
            prior = sum(len(cat_clicks.get(pg, [])) for pg in range(self.current_page))
            page_clicks = cat_clicks.get(self.current_page, [])
            for seq, (pdf_x, pdf_y) in enumerate(page_clicks, prior + 1):
                cx = pdf_x * self.zoom + self.page_offset_x
                cy = pdf_y * self.zoom + self.page_offset_y
                self._draw_marker(cx, cy, seq, cat["color"])

        self._draw_ruler()
        self._update_status()

    def _hex_to_rgb(self, hex_color: str) -> tuple[int, int, int]:
        h = hex_color.lstrip("#")
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

    def _draw_marker(self, cx: float, cy: float, number: int, color: str):
        r    = MARKER_RADIUS
        size = r * 2 + 4        # 44px for radius=20
        rgb  = self._hex_to_rgb(color)

        img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.ellipse([2, 2, size - 2, size - 2],
                     fill=(*rgb, MARKER_ALPHA),
                     outline=(255, 255, 255, 200),
                     width=2)

        # font = self._load_font(max(12, r - 2))
        # try:
        #     draw.text((size / 2, size / 2), str(number),
        #               fill=(255, 255, 255, 255), font=font, anchor="mm")
        # except TypeError:
        #     # Pillow < 8.0 lacks anchor=; fall back to manual centering
        #     bbox = draw.textbbox((0, 0), str(number), font=font)
        #     tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        #     draw.text(((size - tw) / 2, (size - th) / 2), str(number),
        #               fill=(255, 255, 255, 255), font=font)

        photo = ImageTk.PhotoImage(img)
        self._marker_images.append(photo)   # prevent GC
        self.canvas.create_image(cx, cy, image=photo, anchor=tk.CENTER, tags="marker")

    @staticmethod
    def _load_font(size: int):
        candidates = [
            "/System/Library/Fonts/Helvetica.ttc",
            "/System/Library/Fonts/Arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/arial.ttf",
        ]
        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        return ImageFont.load_default()

    # ── Click Handling ────────────────────────────────────────────────────────

    def on_canvas_click(self, event):
        if self.pdf_doc is None:
            return

        if self._pan_key_held:
            self.canvas.scan_mark(event.x, event.y)
            return

        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)

        # Convert canvas → PDF page coordinates
        pdf_x = (cx - self.page_offset_x) / self.zoom
        pdf_y = (cy - self.page_offset_y) / self.zoom

        # ── Ruler mode ────────────────────────────────────────────────────────
        if self.ruler_mode:
            if len(self.ruler_points) >= 2:
                self.ruler_points.clear()   # third click starts a new measurement
            self.ruler_points.append((pdf_x, pdf_y))
            self._refresh_ruler_label()
            self.render_page()
            return

        # Distance guard across ALL categories (avoid stacking markers)
        for cat_idx in range(len(self.categories)):
            for (ex, ey) in self.clicks.get(cat_idx, {}).get(self.current_page, []):
                dist = math.hypot(
                    cx - (ex * self.zoom + self.page_offset_x),
                    cy - (ey * self.zoom + self.page_offset_y),
                )
                if dist < MIN_DISTANCE_PX:
                    self.status_var.set(
                        f"⚠  Too close to an existing marker "
                        f"({dist:.0f}px < {MIN_DISTANCE_PX}px). Ignored."
                    )
                    return

        cat_page = self.clicks.setdefault(self.current_category, {})
        cat_page.setdefault(self.current_page, []).append((pdf_x, pdf_y))

        seq   = sum(len(v) for v in cat_page.values())
        color = self.categories[self.current_category]["color"]
        self._draw_marker(cx, cy, seq, color)
        self._update_status()

    def _update_status(self):
        n_pages    = len(self.pdf_doc) if self.pdf_doc else 0
        page_count = len(self.clicks.get(self.current_category, {}).get(self.current_page, []))
        total      = sum(
            len(pg) for cat in self.clicks.values() for pg in cat.values()
        )
        cat_name = self.categories[self.current_category]["name"]

        self.page_label.config(text=f"Page {self.current_page + 1} / {n_pages}")
        self.count_var.set(f"Total: {total}")
        mode = "📏 RULER" if self.ruler_mode else f"[{cat_name}]"
        self.status_var.set(
            f"{mode}  page count: {page_count}  |  "
            f"all categories total: {total}  |  "
            f"zoom: {self.zoom:.1f}×"
        )

    # ── Controls ──────────────────────────────────────────────────────────────

    def undo_click(self):
        page_clicks = self.clicks.get(self.current_category, {}).get(self.current_page, [])
        if not page_clicks:
            cat_name = self.categories[self.current_category]["name"]
            self.status_var.set(f"Nothing to undo for [{cat_name}] on this page.")
            return
        page_clicks.pop()
        self.render_page()

    def reset_page(self):
        cat_name = self.categories[self.current_category]["name"]
        if not messagebox.askyesno("Reset", f"Clear all [{cat_name}] markers on this page?"):
            return
        self.clicks.setdefault(self.current_category, {})[self.current_page] = []
        self.render_page()

    def prev_page(self):
        if self.pdf_doc and self.current_page > 0:
            self.current_page -= 1
            self.render_page()

    def next_page(self):
        if self.pdf_doc and self.current_page < len(self.pdf_doc) - 1:
            self.current_page += 1
            self.render_page()

    def zoom_in(self):
        self.zoom = min(self.zoom + 0.25, 5.0)
        self.render_page()

    def zoom_out(self):
        self.zoom = max(self.zoom - 0.25, 0.5)
        self.render_page()

    def _on_mousewheel(self, event):
        if event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
        else:
            self.canvas.yview_scroll(1, "units")

    def _on_hscroll(self, event):
        if event.delta > 0:
            self.canvas.xview_scroll(-1, "units")
        else:
            self.canvas.xview_scroll(1, "units")

    def _on_zoom_scroll(self, event):
        if event.num == 4 or event.delta > 0:
            self.zoom_in()
        else:
            self.zoom_out()

    # ── Pan (middle-click drag  or  D + left-click drag) ──────────────────────

    def _set_pan_key(self, held: bool):
        self._pan_key_held = held
        self.canvas.config(cursor="fleur" if held else "crosshair")

    def _pan_start(self, event):
        self.canvas.scan_mark(event.x, event.y)

    def _pan_move(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def _on_b1_motion(self, event):
        if self._pan_key_held:
            self.canvas.scan_dragto(event.x, event.y, gain=1)

    # ── Ruler ─────────────────────────────────────────────────────────────────

    def toggle_ruler_mode(self):
        self.ruler_mode = not self.ruler_mode
        if self.ruler_mode:
            self.ruler_btn.config(relief=tk.SUNKEN, bg="#45475a")
            self.ruler_bar.pack(side=tk.TOP, fill=tk.X)
            self.status_var.set("Ruler mode: click two points to measure distance.")
        else:
            self.ruler_btn.config(relief=tk.FLAT, bg="#313244")
            self.ruler_bar.pack_forget()
            self._update_status()

    def clear_ruler(self):
        self.ruler_points.clear()
        self.ruler_dist_label.config(text="")
        self.render_page()

    # 1 PDF point = 1/72 inch = this many metres on paper
    _PT_TO_M = 0.0254 / 72

    def apply_scale_ratio(self):
        raw = self.scale_ratio_var.get().strip()
        # Accept "200", "1:200", "1/200" — take the part after the separator
        for sep in (":", "/"):
            if sep in raw:
                raw = raw.split(sep)[-1].strip()
                break
        try:
            n = float(raw.replace(",", "."))
            if n <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror(
                "Invalid scale",
                "Enter a positive number for the scale denominator.\n"
                "Example: type  200  for a 1:200 plan.",
                parent=self.root,
            )
            return
        self._apply_scale(n * self._PT_TO_M, ratio=n)

    def set_scale(self):
        if len(self.ruler_points) < 2:
            messagebox.showinfo(
                "Set Scale",
                "Draw a ruler line first:\n"
                "  1. Click 📏 Ruler to enter ruler mode\n"
                "  2. Click two points on the PDF\n"
                "  3. Then click ⚖ Set Scale",
                parent=self.root,
            )
            return
        x1, y1 = self.ruler_points[0]
        x2, y2 = self.ruler_points[1]
        dist_pt = math.hypot(x2 - x1, y2 - y1)
        real_m = simpledialog.askfloat(
            "Set Scale",
            f"Ruler line = {dist_pt:.2f} PDF units.\n"
            f"Enter the real-world length of this line in metres:",
            minvalue=0.0001,
            parent=self.root,
        )
        if real_m:
            m_per_pt = real_m / dist_pt
            equiv_n  = m_per_pt / self._PT_TO_M
            self._apply_scale(m_per_pt, ratio=equiv_n)

    def _apply_scale(self, m_per_pt: float, ratio: float):
        self.scale_m_per_pt = m_per_pt
        self.scale_label.config(text=f"Scale  1:{ratio:.0f}")
        self._refresh_ruler_label()
        self.render_page()

    def _ruler_dist_pdf(self) -> float | None:
        if len(self.ruler_points) < 2:
            return None
        (x1, y1), (x2, y2) = self.ruler_points[0], self.ruler_points[1]
        return math.hypot(x2 - x1, y2 - y1)

    def _refresh_ruler_label(self):
        dist = self._ruler_dist_pdf()
        if dist is None:
            self.ruler_dist_label.config(text="")
            return
        if self.scale_m_per_pt is not None:
            text = f"{dist * self.scale_m_per_pt:.2f} m"
        else:
            text = f"{dist:.2f} pt  (set scale for metres)"
        self.ruler_dist_label.config(text=text)

    def _draw_ruler(self):
        if not self.ruler_points:
            return

        dot_r = 5
        pts_c = [
            (px * self.zoom + self.page_offset_x,
             py * self.zoom + self.page_offset_y)
            for px, py in self.ruler_points
        ]

        # Start endpoint
        x0, y0 = pts_c[0]
        self.canvas.create_oval(x0 - dot_r, y0 - dot_r, x0 + dot_r, y0 + dot_r,
                                fill="#274DEA", outline="#274DEA", width=2, tags="ruler")

        if len(pts_c) == 2:
            x1, y1 = pts_c[1]

            # Line
            self.canvas.create_line(x0, y0, x1, y1,
                                    fill="#274DEA", width=2, dash=(8, 4), tags="ruler")

            # End endpoint
            self.canvas.create_oval(x1 - dot_r, y1 - dot_r, x1 + dot_r, y1 + dot_r,
                                    fill="#274DEA", outline="#274DEA", width=2, tags="ruler")

            # Distance label at midpoint
            mx, my = (x0 + x1) / 2, (y0 + y1) / 2
            dist = self._ruler_dist_pdf()
            if self.scale_m_per_pt is not None:
                label = f" {dist * self.scale_m_per_pt:.2f} m "
            else:
                label = f" {dist:.2f} pt "

            # Background pill behind text
            self.canvas.create_text(mx, my - 15, text=label,
                                    fill="#274DEA", font=("Courier", 10, "bold"),
                                    tags="ruler")

    # ── Export ────────────────────────────────────────────────────────────────

    def export_summary(self):
        has_clicks = any(
            len(pg) > 0
            for cat in self.clicks.values()
            for pg in cat.values()
        )
        if not has_clicks:
            messagebox.showinfo("Export", "No clicks recorded yet.")
            return

        filetypes = [("Text file", "*.txt"), ("JSON file", "*.json")]
        if _XLSX_AVAILABLE:
            filetypes.insert(0, ("Excel workbook", "*.xlsx"))
        filetypes.append(("All files", "*.*"))

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx" if _XLSX_AVAILABLE else ".txt",
            filetypes=filetypes,
            title="Save Summary",
        )
        if not save_path:
            return

        total = sum(len(pg) for cat in self.clicks.values() for pg in cat.values())

        if save_path.endswith(".xlsx"):
            if not _XLSX_AVAILABLE:
                messagebox.showerror("Missing library",
                                     "Install openpyxl first:\n    pip install openpyxl")
                return
            self._export_xlsx(save_path, total)

        elif save_path.endswith(".json"):
            data = {
                "file":  self.pdf_path,
                "total": total,
                "categories": {
                    cat["name"]: {
                        "color": cat["color"],
                        "total": sum(len(v) for v in self.clicks.get(i, {}).values()),
                        "pages": {
                            str(pg + 1): {
                                "count":  len(coords),
                                "clicks": [{"x": round(x, 1), "y": round(y, 1)}
                                           for x, y in coords],
                            }
                            for pg, coords in self.clicks.get(i, {}).items()
                            if coords
                        },
                    }
                    for i, cat in enumerate(self.categories)
                },
            }
            with open(save_path, "w") as f:
                json.dump(data, f, indent=2)

        else:
            lines = [
                "PDF Click Counter — Summary",
                f"File: {self.pdf_path}",
                "",
            ]
            for i, cat in enumerate(self.categories):
                cat_total = sum(len(v) for v in self.clicks.get(i, {}).values())
                lines.append(f"  [{cat['name']}]  colour: {cat['color']}  total: {cat_total}")
                for pg in sorted(self.clicks.get(i, {}).keys()):
                    n = len(self.clicks[i][pg])
                    if n:
                        lines.append(f"    Page {pg + 1}: {n} item(s)")
            lines += ["", f"  GRAND TOTAL: {total} item(s)"]
            with open(save_path, "w") as f:
                f.write("\n".join(lines))

        messagebox.showinfo("Export", f"Summary saved to:\n{save_path}")

    def _export_xlsx(self, path: str, grand_total: int):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Collect all page numbers that have any clicks
        all_pages = sorted({
            pg
            for cat_clicks in self.clicks.values()
            for pg, coords in cat_clicks.items()
            if coords
        })

        # ── Header row (plain text, no fill) ─────────────────────────────────
        headers = ["Category"] + [f"Page {pg + 1}" for pg in all_pages] + ["Total"]
        for col, title in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=title)

        # ── One row per category ──────────────────────────────────────────────
        for row_idx, (i, cat) in enumerate(enumerate(self.categories), 2):
            cat_clicks = self.clicks.get(i, {})
            cat_total  = sum(len(v) for v in cat_clicks.values())

            # Only the category name cell gets the category background colour
            name_cell      = ws.cell(row=row_idx, column=1, value=cat["name"])
            name_cell.fill = PatternFill("solid", fgColor=cat["color"].lstrip("#"))

            for col_idx, pg in enumerate(all_pages, 2):
                ws.cell(row=row_idx, column=col_idx, value=len(cat_clicks.get(pg, [])))

            ws.cell(row=row_idx, column=len(all_pages) + 2, value=cat_total)

        # ── Grand total row (plain text, no fill) ─────────────────────────────
        total_row = len(self.categories) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")

        for col_idx, pg in enumerate(all_pages, 2):
            ws.cell(row=total_row, column=col_idx, value=sum(
                len(self.clicks.get(i, {}).get(pg, []))
                for i in range(len(self.categories))
            ))

        ws.cell(row=total_row, column=len(all_pages) + 2, value=grand_total)

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = max(
            14, max((len(c["name"]) for c in self.categories), default=10) + 2
        )
        for col_idx in range(2, len(all_pages) + 3):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

        wb.save(path)

    # ── Session save / load ───────────────────────────────────────────────────

    def save_session(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".c2c",
            filetypes=[("Click2Count session", "*.c2c"), ("All files", "*.*")],
            title="Save Session",
            initialfile=os.path.splitext(os.path.basename(self.pdf_path))[0] if self.pdf_path else "session",
        )
        if not save_path:
            return

        data = {
            "version":       1,
            "pdf_path":      self.pdf_path,
            "current_page":  self.current_page,
            "zoom":          self.zoom,
            "scale_m_per_pt": self.scale_m_per_pt,
            "categories":    self.categories,
            "clicks": {
                str(cat_idx): {
                    str(pg): coords
                    for pg, coords in pages.items()
                }
                for cat_idx, pages in self.clicks.items()
            },
        }
        with open(save_path, "w") as f:
            json.dump(data, f, indent=2)
        messagebox.showinfo("Session Saved", f"Session saved to:\n{save_path}")

    def load_session(self):
        load_path = filedialog.askopenfilename(
            filetypes=[("Click2Count session", "*.c2c"), ("All files", "*.*")],
            title="Load Session",
        )
        if not load_path:
            return

        try:
            with open(load_path) as f:
                data = json.load(f)
        except Exception as exc:
            messagebox.showerror("Error", f"Could not read session file:\n{exc}")
            return

        pdf_path = data.get("pdf_path", "")
        if pdf_path and not os.path.isfile(pdf_path):
            messagebox.showerror(
                "PDF not found",
                f"The session references a PDF that cannot be found:\n{pdf_path}\n\n"
                "Move the PDF back to its original location and try again.",
            )
            return

        try:
            self.pdf_doc = fitz.open(pdf_path) if pdf_path else None
        except Exception as exc:
            messagebox.showerror("Error", f"Could not open PDF:\n{exc}")
            return

        self.pdf_path      = pdf_path
        self.current_page  = data.get("current_page", 0)
        self.zoom          = data.get("zoom", DEFAULT_ZOOM)
        self.scale_m_per_pt = data.get("scale_m_per_pt")
        self.categories    = data.get("categories", [{"name": "Category 1", "color": "#e63946"}])

        raw_clicks = data.get("clicks", {})
        self.clicks = {
            int(cat_idx): {int(pg): [tuple(xy) for xy in coords] for pg, coords in pages.items()}
            for cat_idx, pages in raw_clicks.items()
        }

        self.current_category = 0
        self._refresh_category_menu()

        if self.scale_m_per_pt is not None:
            ratio = self.scale_m_per_pt / self._PT_TO_M
            self.scale_label.config(text=f"Scale  1:{ratio:.0f}")
            self.scale_ratio_var.set(f"{ratio:.0f}")

        if self.pdf_doc:
            self.root.title(f"PDF Click Counter — {os.path.basename(pdf_path)}")
            self.render_page()


# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app  = PDFClickCounter(root)
    root.mainloop()
